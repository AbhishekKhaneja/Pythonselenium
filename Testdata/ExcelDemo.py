import openpyxl
Book=openpyxl.load_workbook("C:\\Users\\abhkhane\\Documents\\PythonDemo.xlsx")
sheet = Book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
#sheet.cell(row=3,column=5).value="Lovely" #assigning values to cell from script
#print(sheet.cell(row=3,column=5).value)
print(sheet.max_column)
print(sheet.max_row)
dict ={}
#print(sheet['A5'].value)


for i in range(1,sheet.max_row+1):
    if(sheet.cell(row=i,column=1)).value =="TestCase2":
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value


print(dict)